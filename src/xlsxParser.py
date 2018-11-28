from openpyxl import Workbook
from openpyxl import load_workbook
from src.creditEntry import CreditEntry
from src.businessEntry import BusinessEntry
from src.bankEntry import BankEntry
from src.myString import myString
from datetime import datetime
import src.dbAccess as db

def example():
    wb = Workbook()

    # grab the active worksheet
    ws = wb.active

    # Data can be assigned directly to cells
    ws['A1'] = 42

    # Rows can also be appended
    ws.append([1, 2, 3])

    # Python types will automatically be converted
    import datetime
    ws['A2'] = datetime.datetime.now()

    # Save the file
    wb.save("sample.xlsx")

def loadFile(fileName):
    wb = load_workbook(fileName)
    sheets = wb.sheetnames
    print(sheets)
    sheet = wb.active
    return sheet

def process(sheet, date):
    cells = sheet.values

    for row in sheet.iter_rows(min_row=4, min_col=1, max_col=5):
        processRow(row, date)


def handleTotal(total, bankId):
    #Q_POSTPROCESS = 'UPDATE bank_entry SET hide=1 where debit='+total+' and refId='+bankId
    #print(Q_POSTPROCESS)
    #reportData = db.runQuery(Q_POSTPROCESS)
    bList = BankEntry.selectBy(debit=total, refId=bankId)
    for b in bList:
        b.hide = 1


#TODO: handle credit line items
def processRow(row, date):
    for cell in row:
        print(cell.internal_value)

    #date = ;
    #replace = ;
    #refID = ;
    purchaseDate = extractDate(row)
    reportDate = date.strftime("%Y-%m-%d")
    businessName = row[1].internal_value
    amount = extractAmount(row)
    comment = row[4].internal_value
    #entry = BankEntry(date,action,"7872","",amount,"")
    bankId = "8547"

    if(myString.isEmpty(businessName)):
        handleTotal(amount, bankId)
        businessName = "__TOTAL__"

    business = BusinessEntry.selectBy(businessName=businessName).getOne(None)
    if (business == None):
        business = BusinessEntry(businessName=businessName, marketingName=businessName, category="")

    entry = CreditEntry(reportDate=reportDate, purchaseDate=purchaseDate, business=business.id, cardNumber="7872", bankId="8547", credit=0, debit=amount, balance=0)
    entry.toCSV()


def extractDate(row):
    dateString = str(row[0].internal_value)
    if(myString.isEmpty(dateString)):
        return None
    try:
        date = datetime.strptime(dateString, '%Y-%m-%d %H:%M:%S').date().strftime("%Y-%m-%d")
    except:
       date = None
    return date

def extractAmount(row):
    amount = row[3].internal_value
    if(myString.isEmpty(amount)):
            return None
    amount = amount.replace(',','')
    credit = amount[1:len(amount)]
    return credit


#    date = self.extractDate(row)
#    action = self.extractAction(row)
#    refId = self.extractRefId(row)
#    credit = self.extractCredit(row)
#    debit = self.extractDebit(row)
#    balance = self.extractBalance(row)
#    entry = BankEntry(date,action,refId,credit,debit,balance)
#    return entry


def main():
    CreditEntry.createTable(ifNotExists=True)
    BusinessEntry.createTable(ifNotExists=True)
    fileName = "inbox/7872_8547_0205218_Transactions_30_05_2018.xlsx"
    sheet = loadFile(fileName)
    date = datetime(2018,4,2)
    cardNum = "7872"
    bankId = "8547" #replace with the actual value
    process(sheet, date)


main()
