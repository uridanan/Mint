import xlrd
from openpyxl.workbook import Workbook
import datetime
# from openpyxl.utils.cell import get_column_letter
# from openpyxl.reader.excel import load_workbook, InvalidFileException


# Convert XLS to XLSX so it can be parsed
# https://stackoverflow.com/questions/9918646/how-to-convert-xls-to-xlsx

class XLSFile(object):
    data = None

    def __init__(self,fileName):
        self.data = self.loadFile(fileName)

    def getData(self):
        return self.data

    # def loadFile(self, fileName):
    #     # first open using xlrd
    #     book = xlrd.open_workbook(fileName)
    #     index = 0
    #     nrows, ncols = 0, 0
    #     while nrows * ncols == 0:
    #         sheet = book.sheet_by_index(index)
    #         nrows = sheet.nrows
    #         ncols = sheet.ncols
    #         index += 1
    #
    #     # prepare a xlsx sheet
    #     book1 = Workbook()
    #     sheet1 = book1.active
    #
    #     for row in range(1, nrows):
    #         for col in range(1, ncols):
    #             sheet1.cell(row=row, column=col).value = sheet.cell_value(row, col)
    #
    #     return sheet1

    def loadFile(self, fileName):
        #def cvt_xls_to_xlsx(*args, **kw):
        """Open and convert XLS file to openpyxl.workbook.Workbook object
        @param args: args for xlrd.open_workbook
        @param kw: kwargs for xlrd.open_workbook
        @return: openpyxl.workbook.Workbook
        You need -> from openpyxl.utils.cell import get_column_letter
        """

        book_xls = xlrd.open_workbook(fileName, formatting_info=True, ragged_rows=True)
        book_xlsx = Workbook()

        sheet_names = book_xls.sheet_names()
        for sheet_index in range(len(sheet_names)):
            sheet_xls = book_xls.sheet_by_name(sheet_names[sheet_index])

            if sheet_index == 0:
                sheet_xlsx = book_xlsx.active
                sheet_xlsx.title = sheet_names[sheet_index]
            else:
                sheet_xlsx = book_xlsx.create_sheet(title=sheet_names[sheet_index])

            for crange in sheet_xls.merged_cells:
                rlo, rhi, clo, chi = crange

                sheet_xlsx.merge_cells(
                    start_row=rlo + 1, end_row=rhi,
                    start_column=clo + 1, end_column=chi,
                )

            def _get_xlrd_cell_value(cell):
                value = cell.value
                if cell.ctype == xlrd.XL_CELL_DATE:
                    value = datetime.datetime(*xlrd.xldate_as_tuple(value, 0))

                return value

            for row in range(sheet_xls.nrows):
                sheet_xlsx.append((
                    _get_xlrd_cell_value(cell)
                    for cell in sheet_xls.row_slice(row, end_colx=sheet_xls.row_len(row))
                ))

            # for rowx in range(sheet_xls.nrows):
            #     if sheet_xls.rowinfo_map[rowx+1].hidden != 0:
            #         print(sheet_names[sheet_index], rowx)
            #         sheet_xlsx.row_dimensions[rowx+1].hidden = True
            # for coly in range(sheet_xls.ncols):
            #     if sheet_xls.colinfo_map[coly].hidden != 0:
            #         print(sheet_names[sheet_index], coly)
            #         coly_letter = get_column_letter(coly+1)
            #         sheet_xlsx.column_dimensions[coly_letter].hidden = True

        return book_xlsx.active
