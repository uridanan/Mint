from openpyxl import load_workbook

class XSLXFile(object):
    data = None

    def __init__(self,fileName):
        self.data = self.loadFile(fileName)

    def loadFile(self,fileName):
        wb = load_workbook(fileName)
        sheets = wb.sheetnames
        #print(sheets)
        sheet = wb.active
        return sheet

    def getData(self):
        return self.data

